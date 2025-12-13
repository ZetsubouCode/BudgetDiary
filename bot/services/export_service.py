from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from bot.config import EXPORT_DIR
from bot.services.report_service import ReportService
from bot.utils.currency import format_amount
from bot.utils.dates import start_end_of_month


class ExportService:
    def __init__(self, report_service: ReportService):
        self.report = report_service

    async def export_monthly(self, discord_id: int | str, month_start, language: str, region: str) -> Path:
        """Generate an Excel file for the month and return its path."""
        start, end = start_end_of_month(month_start)
        income = await self.report.income_report(discord_id, start, end, language, region)
        outcome = await self.report.outcome_report(discord_id, start, end, language, region)
        balances = await self.report.balances(discord_id, language, region)

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(["Metric", "Value"])
        ws_summary.append(["Total Income", income["total"]])
        ws_summary.append(["Total Outcome", outcome["total"]])

        ws_income = wb.create_sheet("Income")
        ws_income.append(["Date", "Category", "Description", "Amount"])
        for item in income["items"]:
            ws_income.append([item["date"], item["category_name"], item["description"], item["amount"]])

        ws_outcome = wb.create_sheet("Outcome")
        ws_outcome.append(["Date", "Outcome Category", "From Income", "Description", "Amount"])
        for item in outcome["items"]:
            ws_outcome.append(
                [item["date"], item["category_name"], item.get("income_source", ""), item["description"], item["amount"]]
            )

        ws_balance = wb.create_sheet("Balances")
        ws_balance.append(["Income Type", "In", "Out", "Transfer In", "Transfer Out", "Balance"])
        for row in balances:
            ws_balance.append(
                [
                    row["name"],
                    row["in"],
                    row["out"],
                    row["transfer_in"],
                    row["transfer_out"],
                    row["balance"],
                ]
            )

        net = _net_amount(income["total"], outcome["total"], region)
        ws_summary.append(["Net", net])

        filename = EXPORT_DIR / f"{discord_id}_budget_{start:%Y_%m}.xlsx"
        wb.save(filename)
        return filename


def _net_amount(income_total: str, outcome_total: str, region: str) -> str:
    def _to_number(value: str) -> float:
        try:
            number = value.split(" ", 1)[-1].replace(",", "")
            return float(number)
        except Exception:
            return 0.0

    net = _to_number(income_total) - _to_number(outcome_total)
    return format_amount(net, region)
